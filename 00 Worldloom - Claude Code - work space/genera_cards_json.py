"""
genera_cards_json.py

Legge l'Excel di UN mazzo (es. FrostLand_carte.xlsx) e produce il file
cards.json che l'app legge davvero.

Uso da terminale:
    python genera_cards_json.py "Mazzi/Frost Land - Primitivi del ghiaccio/Excel/FrostLand_carte.xlsx"

Non serve saperlo usare da soli: basta chiedere a Claude Code
"rigenera il cards.json di Frost Land dal suo Excel" e lo lancia lui.

Fogli letti:
  - "Carte"      → Pedine, Magie e Trappole (colonna "Tipo Carta" distingue i tre)
  - "Imprevisti" → mazzetto separato (cap. 15 del regolamento), se presente

Regola d'oro: questo script NON si modifica per aggiungere carte.
Le carte si aggiungono SOLO nell'Excel. Lo script si limita a leggere.

Prima di scrivere, lancia tools/validate_cards.py su tutti e quattro gli Excel:
se il validatore trova ERRORI, cards.json NON viene rigenerato. Gli AVVISI non bloccano.
Per saltare il controllo di proposito: --salta-validazione.
"""
import os
import subprocess
import sys
import json
import openpyxl


def intero(valore, default=0):
    try:
        return int(valore)
    except (TypeError, ValueError):
        return default


def intero_o_none(valore):
    """Come intero(), ma restituisce None se la cella e' vuota invece di un default numerico:
    serve per 'Limite Copie', dove vuoto ha un significato preciso (regola standard, cap. editor
    mazzi) diverso da 0 (carta bandita)."""
    try:
        return int(valore)
    except (TypeError, ValueError):
        return None


def sigla(testo):
    return str(testo).strip().lower().replace(" ", "-").replace("'", "")


def nome_file_immagine(nome, variante=1):
    """L'illustrazione dipende da nome + variante, MAI da rarita' o finitura: una stampa Rainbow e
    una Normale della stessa carta condividono lo stesso disegno (il foil e' un effetto, non un file
    diverso). Variante 1 = illustrazione base, nome file invariato; dalla 2 in poi si aggiunge il
    suffisso, es. 'condottiero-fiero-2.png'."""
    suffisso = "" if variante in (None, "", 1) else f"-{variante}"
    return sigla(nome) + suffisso + ".png"


def variante(valore):
    """Colonna 'Varianti Illustrazione': QUALE illustrazione (1 = base). Asse indipendente da
    'Finitura', che dice invece CHE trattamento di stampa ha la stessa illustrazione."""
    try:
        return int(valore)
    except (TypeError, ValueError):
        return 1


def rarita(valore):
    """Colonna 'Rarita'. Cella vuota = 'Comune' (default dichiarato dall'utente il 2026-08-29)."""
    testo = str(valore).strip() if valore is not None else ""
    return testo or "Comune"


def identita(nome, var, rar, fin):
    """Identita' di una carta = Nome + Variante Illustrazione + Rarita' + Finitura (decisione
    dell'utente, 2026-08-29). Serve perche' la stessa carta esiste in piu' STAMPE diverse — la
    Normale e la Rainbow di Condottiero Fiero sono due pezzi da collezione distinti, con la stessa
    illustrazione e le stesse regole. Prima l'unica chiave era il nome: due stampe della stessa
    carta si sarebbero pestate i piedi nel catalogo e nel contatore delle copie dell'editor."""
    return f"{sigla(nome)}__v{var}__{sigla(rar)}__{sigla(fin)}"


def finitura(valore):
    """Colonna 'Finitura' dell'Excel: e' la finitura di STAMPA della carta (Normale, Rainbow, e in
    futuro Star Rail / Restricted / ...), non un si'/no. Viene passata LETTERALE al gioco, che ne
    ricava una classe CSS per valore: cosi' aggiungere una finitura nuova si fa scrivendola nella
    cella, senza toccare ne' questo script ne' il codice React.

    Asse indipendente dalla rarita' (una carta e' p.es. Comune + Rainbow) e da 'Varianti
    Illustrazione' (che invece dice QUALE illustrazione, cioe' un file diverso).
    Cella vuota = 'Normale', cioe' nessun trattamento."""
    testo = str(valore).strip() if valore is not None else ""
    return testo or "Normale"


def leggi_foglio(ws):
    """Restituisce una lista di dizionari {intestazione: valore}, saltando le righe senza nome."""
    intestazioni = [c.value for c in ws[1]]
    righe = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        righe.append({k: v for k, v in zip(intestazioni, row) if k})
    return righe


def genera(percorso_excel):
    wb = openpyxl.load_workbook(percorso_excel, data_only=True)

    alieni, magie, trappole = [], [], []
    for r in leggi_foglio(wb["Carte"]):
        nome = r.get("Nome")
        # "Alieno" e' il vecchio nome del tipo: dal 2026-08-29 nel gioco esistono solo Pedine
        # (rinomina terminologica richiesta dall'utente, applicata insieme agli Excel e al motore).
        # Gli Excel vecchi e i backup dicono ancora "Alieno": accettarlo come sinonimo evita che una
        # riga non aggiornata sparisca in silenzio dal mazzo, che e' esattamente il bug che le 15
        # righe gia' marcate "Pedina" avevano prima di questa modifica.
        tipo_carta = (r.get("Tipo Carta") or "Pedina").strip()
        if tipo_carta == "Alieno":
            tipo_carta = "Pedina"
        effetto = {
            "tipo": r.get("Tipo Effetto") or "none",
            "testo": r.get("Testo Effetto") or None,
            "codice": r.get("Codice Effetto (per Claude Code)") or None,
        }
        var = variante(r.get("Varianti Illustrazione"))
        rar = rarita(r.get("Rarita"))
        fin = finitura(r.get("Finitura"))
        base = {
            "id": identita(nome, var, rar, fin),
            "nome": nome,
            "variante": var,
            "rarita": rar,
            "copie": intero(r.get("Copie"), 1),
            # Limite di costruzione mazzo (cap. editor mazzi): None = regola standard (max 3, o
            # 'copie' se inferiore), un numero = eccezione esplicita, 0 = carta bandita.
            "limiteCopie": intero_o_none(r.get("Limite Copie")),
            "finitura": fin,
            "effetto": effetto,
        }

        if tipo_carta == "Pedina":
            alieni.append({
                **base,
                "tipoCarta": "pedina",
                "archetipo": r.get("Archetipo"),
                "livello": intero(r.get("Livello"), 1),
                "ruolo": r.get("Ruolo"),
                "vita": intero(r.get("Vita")),
                "attacco": intero(r.get("Attacco")),
                "parata": intero(r.get("Parata")),
                "attacchi": intero(r.get("Attacchi"), 2),
                "immagine": nome_file_immagine(nome, var),
            })
        elif tipo_carta == "Magia":
            # Il sottotipo di gioco (cap. 14: normale/terreno/continua/rapida) NON è la colonna
            # "Tipo Effetto" (quella è solo una categoria descrittiva libera, es. "kill", "copy").
            # Nessuna colonna Excel dedicata esiste ancora per continua/rapida: per ora si deduce
            # "terreno" dal codice effetto (prefisso terr_), altrimenti "normale".
            codice_effetto = effetto["codice"] or ""
            sottotipo = "terreno" if codice_effetto.startswith("terr_") else "normale"
            magie.append({**base, "tipoCarta": "magia", "sottotipo": sottotipo})
        elif tipo_carta == "Trappola":
            trappole.append({**base, "tipoCarta": "trappola"})

    imprevisti = []
    if "Imprevisti" in wb.sheetnames:
        for r in leggi_foglio(wb["Imprevisti"]):
            nome_imp = r.get("Nome")
            var_imp = variante(r.get("Varianti Illustrazione"))
            rar_imp = rarita(r.get("Rarita"))
            fin_imp = finitura(r.get("Finitura"))
            imprevisti.append({
                "id": identita(nome_imp, var_imp, rar_imp, fin_imp),
                "nome": nome_imp,
                "variante": var_imp,
                "rarita": rar_imp,
                "tipoCarta": "imprevisto",
                "copie": intero(r.get("Copie"), 1),
                # Stessa regola del Worldloom, ma la regola standard di default e' max 2 (cap. 15
                # del regolamento: "massimo 2 copie identiche" nel mazzetto Imprevisti), non max 3.
                "limiteCopie": intero_o_none(r.get("Limite Copie")),
                "finitura": fin_imp,
                "effetto": {
                    "tipo": "imprevisto",
                    "testo": r.get("Testo Effetto") or None,
                    "codice": r.get("Codice Effetto (per Claude Code)") or None,
                },
            })

    worldloom = alieni + magie + trappole
    out = {
        "generato_da": percorso_excel,
        "numero_carte_uniche": len(worldloom),
        "numero_copie_totali": sum(c["copie"] for c in worldloom),
        # "carte" resta il Worldloom (mazzo principale): alieni + magie + trappole
        "carte": worldloom,
        # gli Imprevisti sono un mazzetto separato (cap. 15)
        "imprevisti": imprevisti,
    }

    cartella_mazzo = os.path.dirname(os.path.dirname(os.path.abspath(percorso_excel)))
    percorso_output = os.path.join(cartella_mazzo, "cards.json")
    with open(percorso_output, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    # L'identita' (nome+variante+rarita'+finitura) deve essere unica: due righe identiche su tutti e
    # quattro i campi sarebbero due carte indistinguibili per il motore. Segnalato, non corretto in
    # silenzio: e' un errore di compilazione dell'Excel, va risolto la'.
    visti = {}
    for c in worldloom + imprevisti:
        visti.setdefault(c["id"], []).append(c["nome"])
    doppi = {k: v for k, v in visti.items() if len(v) > 1}
    if doppi:
        print(f"ATTENZIONE: {len(doppi)} identita' duplicate (stesso nome+variante+rarita+finitura):")
        for k, v in doppi.items():
            print(f"  {k}  x{len(v)}")

    print(f"OK: {len(alieni)} Pedine, {len(magie)} Magie, {len(trappole)} Trappole, {len(imprevisti)} Imprevisti")
    print(f"    Worldloom: {out['numero_copie_totali']} copie totali")
    print(f"Scritto in: {percorso_output}")
    print()
    print("Immagini attese in Images/ (solo per le Pedine):")
    for c in alieni:
        print(f"  {c['nome']:30s} -> Images/{c['immagine']}")


def cancello_validazione():
    """Gate: cards.json non si rigenera se tools/validate_cards.py trova ERRORI.

    Gli AVVISI non bloccano (sono i punti in attesa di una decisione dell'utente).
    Se il validatore non e' installabile/eseguibile lo si dice e si prosegue: meglio un
    avviso rumoroso che una pipeline che si pianta per un motivo estraneo alle carte.
    Per saltarlo di proposito: --salta-validazione.
    """
    if "--salta-validazione" in sys.argv:
        print("!! validazione SALTATA su richiesta (--salta-validazione)")
        return
    percorso = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tools", "validate_cards.py")
    if not os.path.exists(percorso):
        print(f"!! validatore non trovato ({percorso}): rigenerazione senza controlli")
        return
    print("Validazione dati carte…")
    esito = subprocess.run([sys.executable, percorso, "--solo-errori"])
    if esito.returncode != 0:
        print()
        print("STOP: il validatore ha trovato errori. cards.json NON e' stato rigenerato.")
        print("      Correggi l'Excel, poi rilancia. Report completo: python tools/validate_cards.py")
        sys.exit(esito.returncode)


if __name__ == "__main__":
    argomenti = [a for a in sys.argv[1:] if not a.startswith("--")]
    if len(argomenti) != 1:
        print("Uso: python genera_cards_json.py percorso/al/Excel/NomeMazzo_carte.xlsx [--salta-validazione]")
        sys.exit(1)
    cancello_validazione()
    genera(argomenti[0])

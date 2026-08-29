# -*- coding: utf-8 -*-
"""Sync dei materiali carte dal progetto verso il sito.

REGOLA: i dati delle carte non si scrivono mai a mano nel sito. Si rigenerano da qui.

Fonti (in ordine di autorità):
  1. Gli Excel dei mazzi  -> elenco delle stampe, Finitura, Rarita, Autore, Varianti.
     L'Excel ha UNA RIGA PER STAMPA: la stessa carta compare una volta come "Normale"
     e una volta come "Rainbow", su righe adiacenti. È da qui che il sito sa quali
     carte sono foil.
  2. I cards.json generati -> statistiche di gioco ed effetti (vita, attacco, ...).
  3. Le "Complete cards compressed" -> le immagini (Normale e Rainbow condividono
     la stessa illustrazione: il foil sul sito è un effetto, non un file diverso).

Uso:  python "Sito web - Social/worldloomtcg/sync_carte.py"
"""
import json
import shutil
import unicodedata
from datetime import date
from pathlib import Path

import openpyxl

QUI = Path(__file__).resolve().parent
RADICE = QUI.parent.parent

MAZZI = [
    {
        "id": "frost-land",
        "nome": "Frost Land",
        "sottotitolo": "Primitivi del ghiaccio",
        "excel": RADICE / "Mazzi/Frost Land - Primitivi del ghiaccio/Excel/FrostLand_carte.xlsx",
        "dati": RADICE / "App - HTML - Test/src/data/generated/mazzi/frost-land/cards.json",
        "immagini": RADICE / "Mazzi/Frost Land - Primitivi del ghiaccio/Complete cards compressed",
    },
    {
        "id": "kepler-452b",
        "nome": "Kepler-452B",
        "sottotitolo": "Manipolatrici d'aura",
        "excel": RADICE / "Mazzi/Marbion - Kepler - 452 B - Manipolatrici d'aura/Excel/Kepler452B_carte.xlsx",
        "dati": RADICE / "App - HTML - Test/src/data/generated/mazzi/kepler-452b/cards.json",
        "immagini": RADICE / "Mazzi/Marbion - Kepler - 452 B - Manipolatrici d'aura/Complete cards compressed",
    },
]

DEST_IMG = QUI / "assets/img/carte"
DEST_DATI = QUI / "assets/data/carte.json"

# Risorse di marca copiate dall'app, per non tenerne due copie divergenti
RISORSE = [
    (RADICE / "App - HTML - Test/src/assets/pittogramma.png", QUI / "assets/img/pittogramma.png"),
    (RADICE / "App - HTML - Test/src/assets/logo-testo.png", QUI / "assets/img/logo-testo.png"),
    (RADICE / "App - HTML - Test/src/assets/fonts/cinzel.ttf", QUI / "assets/fonts/cinzel.ttf"),
    (RADICE / "App - HTML - Test/src/assets/fonts/cormorant-garamond.ttf", QUI / "assets/fonts/cormorant-garamond.ttf"),
    (RADICE / "App - HTML - Test/src/assets/fonts/cinzel-OFL.txt", QUI / "assets/fonts/cinzel-OFL.txt"),
    (RADICE / "App - HTML - Test/src/assets/fonts/cormorant-garamond-OFL.txt", QUI / "assets/fonts/cormorant-garamond-OFL.txt"),
]

# Quali valori della colonna Finitura contano come "foil" (effetto pieno sul sito)
FINITURE_FOIL = {"rainbow", "foil", "holo", "olografica"}


def nome_file_output(nome_carta):
    """Stessa identica regola di componi_carte.py (nome_file_output, riga 288)."""
    return nome_carta.lower().replace(" ", "-").replace("'", "").replace("’", "") + ".jpg"


def chiave(nome):
    """Confronto tollerante fra Excel e cards.json: gli Excel scrivono a volte
    'e'' al posto di 'è'. Si confronta senza accenti, punteggiatura e maiuscole."""
    n = unicodedata.normalize("NFKD", str(nome or ""))
    n = "".join(c for c in n if not unicodedata.combining(c))
    return "".join(c for c in n.lower() if c.isalnum())


def testo(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def main():
    carte = []
    senza_immagine = []
    DEST_IMG.mkdir(parents=True, exist_ok=True)

    for mazzo in MAZZI:
        # --- statistiche di gioco dal cards.json, indicizzate per nome normalizzato
        stats = {}
        if mazzo["dati"].exists():
            dati = json.loads(mazzo["dati"].read_text(encoding="utf-8"))
            for c in dati["carte"]:
                stats[chiave(c["nome"])] = c

        # --- elenco delle stampe dall'Excel
        ws = openpyxl.load_workbook(mazzo["excel"], data_only=True)["Carte"]
        intest = [str(c.value) for c in ws[1]]
        col = {h: i + 1 for i, h in enumerate(intest)}

        def val(riga, nome_col):
            return testo(ws.cell(riga, col[nome_col]).value) if nome_col in col else None

        dest_mazzo = DEST_IMG / mazzo["id"]
        dest_mazzo.mkdir(parents=True, exist_ok=True)
        elenco_file = {p.name for p in mazzo["immagini"].iterdir()} if mazzo["immagini"].exists() else set()

        for riga in range(2, ws.max_row + 1):
            nome = val(riga, "Nome")
            if not nome:
                continue

            finitura = val(riga, "Finitura") or "Normale"
            e_foil = finitura.lower() in FINITURE_FOIL

            # Normale e Rainbow condividono l'illustrazione: il foil è un effetto CSS
            file_img = nome_file_output(nome)
            if file_img in elenco_file:
                shutil.copyfile(mazzo["immagini"] / file_img, dest_mazzo / file_img)
                percorso = f"assets/img/carte/{mazzo['id']}/{file_img}"
            else:
                percorso = None
                senza_immagine.append(f"{mazzo['nome']} · {nome} ({finitura})")

            s = stats.get(chiave(nome), {})
            carte.append(
                {
                    "mazzo": mazzo["id"],
                    "mazzoNome": mazzo["nome"],
                    "nome": nome,
                    "immagine": percorso,
                    # --- dalla riga di Excel: identificano LA STAMPA
                    "finitura": finitura,
                    "foil": e_foil,
                    "variante": val(riga, "Varianti Illustrazione"),
                    "rarita": val(riga, "Rarita"),
                    "autore": val(riga, "Autore"),
                    # --- dal cards.json: descrivono LA CARTA
                    "tipoCarta": s.get("tipoCarta") or val(riga, "Tipo Carta"),
                    "sottotipo": s.get("sottotipo") or val(riga, "Sottotipo"),
                    "archetipo": (s.get("archetipo") or val(riga, "Archetipo") or "").rstrip(", ") or None,
                    "livello": s.get("livello"),
                    "ruolo": s.get("ruolo") or val(riga, "Ruolo"),
                    "vita": s.get("vita"),
                    "attacco": s.get("attacco"),
                    "parata": s.get("parata"),
                    "attacchi": s.get("attacchi"),
                    "copie": s.get("copie"),
                    "effetto": (s.get("effetto") or {}).get("testo") or val(riga, "Testo Effetto"),
                }
            )

    # L'ordine è quello dell'Excel, quindi le finiture della stessa carta restano
    # adiacenti (prima Normale, poi Rainbow), come richiesto.

    for sorgente, destinazione in RISORSE:
        if sorgente.exists():
            destinazione.parent.mkdir(parents=True, exist_ok=True)
            shutil.copyfile(sorgente, destinazione)

    DEST_DATI.parent.mkdir(parents=True, exist_ok=True)
    DEST_DATI.write_text(
        json.dumps(
            {
                "generato": date.today().isoformat(),
                "avvertenza": "File generato da sync_carte.py. Non modificare a mano.",
                "mazzi": [{k: m[k] for k in ("id", "nome", "sottotitolo")} for m in MAZZI],
                "carte": carte,
            },
            ensure_ascii=False,
            indent=1,
        ),
        encoding="utf-8",
    )

    foil = sum(1 for c in carte if c["foil"])
    print(f"OK {len(carte)} stampe scritte in assets/data/carte.json")
    print(f"   di cui {foil} foil e {len(carte) - foil} normali")
    if senza_immagine:
        print(f"\n!! {len(senza_immagine)} stampe senza immagine:")
        for m in senza_immagine:
            print("   - " + m)


if __name__ == "__main__":
    main()

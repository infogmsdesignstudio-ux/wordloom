"""
validate_cards.py — controllo di integrita' dei dati carte di Worldloom.

Gira su TUTTI e quattro gli Excel (i due ufficiali *_carte.xlsx e i due *_proposte.xlsx)
ed esce con codice != 0 se trova almeno un ERRORE. Gli AVVISI non fanno fallire il gate:
sono cose gia' note e in attesa di una decisione dell'utente.

Uso:
    python tools/validate_cards.py                 # tutti e quattro gli Excel
    python tools/validate_cards.py --solo-errori   # nasconde gli avvisi
    python tools/validate_cards.py --range-strict  # i range statistiche diventano ERRORI
    python tools/validate_cards.py --json          # report leggibile da altri script

Aggancio al build: genera_cards_json.py lo chiama prima di scrivere cards.json.
Se il validatore fallisce, cards.json NON viene rigenerato.

I vocabolari chiusi stanno in tools/vocabolari.json, il glossario keyword in tools/keywords.json.
I range delle statistiche NON sono scritti qui: vengono LETTI dal cap. 8 del
Regolamento/Worldloom_Regolamento_v2.1.html a ogni esecuzione.
"""
import argparse
import glob
import json
import os
import re
import sys
import unicodedata

import openpyxl

# La console di Windows e' cp1252: senza questo, le frecce e le virgolette del report
# fanno esplodere lo script invece di stamparlo.
for _flusso in (sys.stdout, sys.stderr):
    try:
        _flusso.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

RADICE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CARTELLA_TOOLS = os.path.join(RADICE, "tools")
REGOLAMENTO = os.path.join(RADICE, "Regolamento", "Worldloom_Regolamento_v2.1.html")

ERRORE = "ERRORE"
AVVISO = "AVVISO"


# ─────────────────────────────────────────────────────────────────────────────
# Raccolta dei dati
# ─────────────────────────────────────────────────────────────────────────────

class Riga:
    """Una riga di un foglio Excel, con il suo indirizzo completo per il report."""

    def __init__(self, file_path, foglio, numero_riga, dati):
        self.file = os.path.basename(file_path)
        self.file_path = file_path
        self.foglio = foglio
        self.riga = numero_riga
        self.dati = dati
        self.ufficiale = "_carte" in self.file

    def get(self, colonna, default=None):
        v = self.dati.get(colonna, default)
        if isinstance(v, str):
            v = v.strip()
        return v if v not in ("", None) else default

    @property
    def nome(self):
        return self.get("Nome") or "(senza nome)"

    @property
    def tipo_carta(self):
        t = self.get("Tipo Carta") or ("Imprevisto" if self.foglio == "Imprevisti" else "Pedina")
        return "Pedina" if t == "Alieno" else t

    @property
    def e_pedina(self):
        return self.foglio == "Carte" and self.tipo_carta == "Pedina"

    def __str__(self):
        return f"{self.file} · {self.foglio} · r{self.riga} · «{self.nome}»"


def trova_excel():
    percorsi = sorted(glob.glob(os.path.join(RADICE, "Mazzi", "*", "Excel", "*.xlsx")))
    if not percorsi:
        raise SystemExit("Nessun Excel trovato sotto Mazzi/*/Excel/. Controlla il percorso del progetto.")
    return percorsi


def leggi_righe(percorso):
    wb = openpyxl.load_workbook(percorso, data_only=True)
    righe = []
    for foglio in ("Carte", "Imprevisti"):
        if foglio not in wb.sheetnames:
            continue
        ws = wb[foglio]
        intestazioni = [c.value for c in ws[1]]
        for i, valori in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not valori or valori[0] is None or str(valori[0]).strip() == "":
                continue
            dati = {k: v for k, v in zip(intestazioni, valori) if k}
            righe.append(Riga(percorso, foglio, i, dati))
    return righe


# ─────────────────────────────────────────────────────────────────────────────
# Range statistiche: LETTI dal regolamento, non assunti
# ─────────────────────────────────────────────────────────────────────────────

def leggi_range_regolamento(percorso_html=REGOLAMENTO):
    """Estrae la tabella dei range dal cap. 8 ('I range delle statistiche').

    Struttura attesa:
        <tr><th>Livello</th><th>Attacco e Parata</th><th>Vita</th><th>Note</th></tr>
        <tr><td>1</td><td>1 – 12</td><td>6 – 16</td>…
    Se la tabella cambia forma, questa funzione fallisce rumorosamente invece di
    inventare dei numeri.
    """
    if not os.path.exists(percorso_html):
        raise SystemExit(f"Regolamento non trovato: {percorso_html}")
    with open(percorso_html, encoding="utf-8") as f:
        html = f.read()

    m = re.search(r'<section id="c8">(.*?)</section>', html, re.S)
    if not m:
        raise SystemExit("cap. 8 non trovato nel regolamento (section id='c8'): controlla il file.")
    sezione = m.group(1)

    tabelle = re.findall(r"<table>(.*?)</table>", sezione, re.S)
    if not tabelle:
        raise SystemExit("cap. 8: nessuna tabella trovata.")

    range_per_livello = {}
    for righe_html in re.findall(r"<tr>(.*?)</tr>", tabelle[0], re.S):
        celle = [re.sub(r"<[^>]+>", "", c).strip() for c in re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", righe_html, re.S)]
        if len(celle) < 3:
            continue
        liv = re.sub(r"\D", "", celle[0])
        if not liv:
            continue
        # "1 – 12" (trattino lungo o corto)
        def coppia(testo):
            n = re.findall(r"\d+", testo)
            return (int(n[0]), int(n[1])) if len(n) >= 2 else None

        ap, vita = coppia(celle[1]), coppia(celle[2])
        if ap and vita:
            range_per_livello[int(liv)] = {"attacco_parata": ap, "vita": vita}

    if not range_per_livello:
        raise SystemExit("cap. 8: tabella dei range illeggibile (nessun livello estratto).")
    return range_per_livello


# ─────────────────────────────────────────────────────────────────────────────
# Budget di testo: misurato con la STESSA funzione di componi_carte.py
# ─────────────────────────────────────────────────────────────────────────────

def costruisci_misuratore(budget):
    """Ritorna (conta_righe, descrizione). Usa PIL e i font reali se disponibili,
    altrimenti ritorna (None, motivo) e il controllo 9 viene saltato dichiarandolo."""
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        return None, "Pillow non installato"

    percorso_font = os.path.join("C:/Windows/Fonts", budget["pedina"]["font"])
    if not os.path.exists(percorso_font):
        return None, f"font non trovato: {percorso_font}"

    draw = ImageDraw.Draw(Image.new("RGB", (10, 10)))
    cache = {}

    def conta_righe(testo, dimensione_pt, larghezza_px):
        """Replica esatta dell'andare a capo di componi_carte.py:testo_a_capo()."""
        if not testo:
            return 0
        if dimensione_pt not in cache:
            cache[dimensione_pt] = ImageFont.truetype(percorso_font, dimensione_pt)
        font_obj = cache[dimensione_pt]
        righe, riga = 0, ""
        for parola in str(testo).split():
            prova = (riga + " " + parola).strip()
            if draw.textlength(prova, font=font_obj) > larghezza_px and riga:
                righe += 1
                riga = parola
            else:
                riga = prova
        if riga:
            righe += 1
        return righe

    return conta_righe, f"PIL + {os.path.basename(percorso_font)}"


# ─────────────────────────────────────────────────────────────────────────────
# Utility
# ─────────────────────────────────────────────────────────────────────────────

def normalizza(testo):
    """Per confrontare due testi 'uguali': niente accenti, niente doppi spazi, minuscolo,
    apostrofi tipografici uniformati."""
    if testo is None:
        return ""
    t = str(testo).replace("\u2019", "'").replace("\u2018", "'")
    t = unicodedata.normalize("NFKD", t)
    t = "".join(c for c in t if not unicodedata.combining(c))
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t.rstrip(".")


def carica_json(nome):
    percorso = os.path.join(CARTELLA_TOOLS, nome)
    if not os.path.exists(percorso):
        raise SystemExit(f"File di configurazione mancante: {percorso}")
    with open(percorso, encoding="utf-8") as f:
        return json.load(f)


# ─────────────────────────────────────────────────────────────────────────────
# I controlli
# ─────────────────────────────────────────────────────────────────────────────

class Report:
    def __init__(self):
        self.voci = []

    def aggiungi(self, livello, controllo, riga, messaggio):
        self.voci.append({
            "livello": livello,
            "controllo": controllo,
            "file": riga.file if riga else "-",
            "foglio": riga.foglio if riga else "-",
            "riga": riga.riga if riga else None,
            "carta": riga.nome if riga else "-",
            "messaggio": messaggio,
        })

    @property
    def errori(self):
        return [v for v in self.voci if v["livello"] == ERRORE]

    @property
    def avvisi(self):
        return [v for v in self.voci if v["livello"] == AVVISO]


def controllo_1_tipo_effetto(righe, vocab, rep):
    ammessi = set(vocab["tipo_effetto"]["ammessi"])
    proposta = vocab["tipo_effetto"]["proposta_lista_chiusa"]
    mappa = dict(proposta["mappa_conversione"]["_certe"])
    da_confermare = proposta["mappa_conversione"]["_da_confermare_con_utente"]

    for r in righe:
        if r.foglio == "Imprevisti":
            continue  # gli Imprevisti non hanno la colonna Tipo Effetto
        valore = r.get("Tipo Effetto")
        if valore is None:
            rep.aggiungi(ERRORE, "1·tipo-effetto", r, "colonna 'Tipo Effetto' vuota")
            continue
        if valore not in ammessi:
            rep.aggiungi(ERRORE, "1·tipo-effetto", r,
                         f"valore «{valore}» fuori dal vocabolario chiuso di tools/vocabolari.json")
            continue
        if valore in mappa:
            rep.aggiungi(AVVISO, "1·tipo-effetto·consolidamento", r,
                         f"«{valore}» → «{mappa[valore]}» secondo la proposta di lista chiusa (NON applicata, serve ok utente)")
        elif valore in da_confermare:
            candidati = " / ".join(da_confermare[valore]["candidati"])
            rep.aggiungi(AVVISO, "1·tipo-effetto·da-decidere", r,
                         f"«{valore}» non ha una destinazione univoca nella proposta (candidati: {candidati}) — serve una scelta dell'utente")


def controllo_2_ruolo(righe, vocab, rep):
    ammessi = set(vocab["ruolo"]["ammessi"])
    for r in righe:
        if not r.e_pedina:
            continue
        valore = r.get("Ruolo")
        if valore is None:
            rep.aggiungi(ERRORE, "2·ruolo", r, "colonna 'Ruolo' vuota su una Pedina")
        elif valore not in ammessi:
            rep.aggiungi(ERRORE, "2·ruolo", r,
                         f"ruolo «{valore}» non ammesso (ammessi: {', '.join(sorted(ammessi))})")


def controllo_3_archetipo(righe, vocab, rep):
    ammessi = set(vocab["archetipo"]["ammessi"])
    for r in righe:
        if not r.e_pedina:
            continue
        valore = r.get("Archetipo")
        if valore is None:
            rep.aggiungi(ERRORE, "3·archetipo", r, "colonna 'Archetipo' vuota su una Pedina")
        elif valore not in ammessi:
            rep.aggiungi(ERRORE, "3·archetipo", r,
                         f"archetipo «{valore}» non nella Ruota (ammessi: {', '.join(ammessi)})")


def controllo_4_range(righe, range_liv, rep, severita):
    for r in righe:
        if not r.e_pedina:
            continue
        try:
            livello = int(r.get("Livello"))
        except (TypeError, ValueError):
            rep.aggiungi(ERRORE, "4·range", r, "colonna 'Livello' vuota o non numerica su una Pedina")
            continue
        if livello not in range_liv:
            rep.aggiungi(ERRORE, "4·range", r, f"Livello {livello} non previsto dal cap. 8 del regolamento")
            continue
        banda = range_liv[livello]
        for etichetta, colonna, chiave in (
            ("Vita", "Vita", "vita"),
            ("Attacco", "Attacco", "attacco_parata"),
            ("Parata", "Parata", "attacco_parata"),
        ):
            try:
                valore = int(r.get(colonna))
            except (TypeError, ValueError):
                rep.aggiungi(ERRORE, "4·range", r, f"colonna '{colonna}' vuota o non numerica")
                continue
            lo, hi = banda[chiave]
            if not (lo <= valore <= hi):
                rep.aggiungi(severita, "4·range", r,
                             f"L{livello} · {etichetta}={valore} fuori dal range {lo}–{hi} (cap. 8)")


def controllo_5_codici_effetto(righe, rep):
    per_testo, per_codice = {}, {}
    for r in righe:
        testo = normalizza(r.get("Testo Effetto"))
        codice = r.get("Codice Effetto (per Claude Code)")
        if not testo or not codice:
            continue
        per_testo.setdefault(testo, {}).setdefault(codice, []).append(r)
        per_codice.setdefault(codice, {}).setdefault(testo, []).append(r)

    for testo, codici in per_testo.items():
        if len(codici) > 1:
            elenco = []
            for codice, rr in sorted(codici.items()):
                elenco.append(f"{codice} ({', '.join(sorted({x.nome for x in rr}))})")
            prima = sorted(codici.values(), key=lambda rr: (rr[0].file, rr[0].riga))[0][0]
            rep.aggiungi(AVVISO, "5·testo-uguale-codici-diversi", prima,
                         "stesso Testo Effetto parola per parola, codici diversi: " + " · ".join(elenco))

    for codice, testi in per_codice.items():
        if len(testi) > 1:
            prima = sorted(testi.values(), key=lambda rr: (rr[0].file, rr[0].riga))[0][0]
            varianti = " ||| ".join(sorted(t[:90] for t in testi))
            rep.aggiungi(ERRORE, "5·codice-uguale-testi-diversi", prima,
                         f"codice «{codice}» usato con {len(testi)} testi diversi: {varianti}")


def controllo_6_campi_stampa(righe, vocab, rep):
    campi = vocab["campi_stampa_obbligatori"]["campi"]
    mancanti = {c: [] for c in campi}
    for r in righe:
        if not r.ufficiale:
            continue  # le proposte non vanno in stampa
        for c in campi:
            if c not in r.dati:
                continue  # la colonna non esiste in questo foglio: non e' un buco di compilazione
            if r.get(c) is None:
                mancanti[c].append(r)
    for c, rr in mancanti.items():
        if rr:
            files = sorted({x.file for x in rr})
            rep.aggiungi(AVVISO, "6·campi-stampa", None,
                         f"colonna '{c}' vuota su {len(rr)} righe ufficiali ({', '.join(files)}) — "
                         f"warning per scelta esplicita: si promuove a errore quando l'utente li compila")


RE_TOKEN = re.compile(r"[A-ZÀ-ÖØ-Þ][\wÀ-ÖØ-öø-ÿ']{2,}")


def controllo_7_keyword_orfane(righe, glossario, rep):
    ammesse = set()
    for sezione in glossario.values():
        if isinstance(sezione, dict) and isinstance(sezione.get("termini"), list):
            ammesse.update(sezione["termini"])

    # ogni parola che compare dentro il NOME di una carta esistente non e' una keyword di regola
    parole_dei_nomi = set()
    for r in righe:
        for p in RE_TOKEN.findall(r.nome):
            parole_dei_nomi.add(p)

    for r in righe:
        testo = r.get("Testo Effetto")
        if not testo:
            continue
        gia_segnalate = set()
        # spezza in frasi per poter ignorare le maiuscole di inizio frase
        for frase in re.split(r"(?<=[.:;!?])\s+|\n|(?<= - )", str(testo)):
            frase = frase.strip()
            if not frase:
                continue
            for m in RE_TOKEN.finditer(frase):
                if m.start() == 0:
                    continue  # inizio frase: la maiuscola non significa niente
                # gli apici che delimitano una citazione ("con 'Goblin' nel nome") non fanno parte
                # della parola
                parola = m.group(0).strip("'’")
                if not parola or parola in ammesse or parola in parole_dei_nomi:
                    continue
                if parola.isupper():
                    continue  # etichette tipo TERRENO / TRAPPOLA / MAGIA RAPIDA
                if parola in gia_segnalate:
                    continue  # una segnalazione per parola per carta, non una per occorrenza
                gia_segnalate.add(parola)
                rep.aggiungi(AVVISO, "7·keyword-orfana", r,
                             f"parola chiave «{parola}» non definita in tools/keywords.json ne' in un nome di carta")


def controllo_8_doppia_presenza(righe, rep):
    ufficiali = {}
    for r in righe:
        if r.ufficiale:
            ufficiali.setdefault((normalizza(r.nome), r.foglio), r)
    for r in righe:
        if r.ufficiale:
            continue
        gemella = ufficiali.get((normalizza(r.nome), r.foglio))
        if gemella:
            rep.aggiungi(ERRORE, "8·doppia-presenza", r,
                         f"presente anche nell'Excel ufficiale ({gemella.file} · {gemella.foglio} · r{gemella.riga}): "
                         f"una carta gia' in gioco non puo' restare tra le proposte")


def controllo_9_budget_testo(righe, vocab, rep):
    budget = vocab["budget_testo"]
    conta_righe, come = costruisci_misuratore(budget)
    if conta_righe is None:
        rep.aggiungi(AVVISO, "9·budget-testo", None,
                     f"controllo SALTATO ({come}): impossibile misurare l'andata a capo reale. "
                     f"Nessun testo e' stato verificato.")
        return come

    for r in righe:
        testo = r.get("Testo Effetto")
        if not testo:
            continue
        b = budget["pedina"] if r.e_pedina else budget["magia_trappola_terreno_imprevisto"]
        n = conta_righe(testo, b["dimensione_pt"], b["larghezza_px"])
        if n > b["max_righe"]:
            rep.aggiungi(AVVISO, "9·budget-testo·componi_carte", r,
                         f"{n} righe contro {b['max_righe']} disponibili: componi_carte.py TRONCA il testo "
                         f"con '…' sulla carta composta (eccedenza: {n - b['max_righe']} righe)")

    # Il secondo renderer (worldloom-cards/render.js) dichiara un budget diverso in SPEC.md §6
    # e ha un rilevatore di overflow proprio: non lo si duplica qui a occhio.
    contraddizione = budget.get("contraddizione_da_risolvere")
    if contraddizione:
        rep.aggiungi(AVVISO, "9·budget-testo·due-renderer", None,
                     "esistono DUE renderer con budget diversi (componi_carte.py 3/6 righe · "
                     "worldloom-cards/render.js ~6/~9 secondo SPEC.md §6) e i documenti non concordano "
                     "su quale sia vivo (Roadmap T.9 vs CLAUDE.md/skill pipeline-carte). "
                     "Il controllo sopra misura solo componi_carte.py; per l'altro: `node render.js` "
                     "in Mazzi/00 Layout generico/worldloom-cards/ stampa da solo le carte in overflow.")
    return come


# ─────────────────────────────────────────────────────────────────────────────
# Esecuzione
# ─────────────────────────────────────────────────────────────────────────────

def valida(range_strict=False):
    vocab = carica_json("vocabolari.json")
    glossario = carica_json("keywords.json")
    range_liv = leggi_range_regolamento()

    percorsi = trova_excel()
    righe = []
    for p in percorsi:
        righe.extend(leggi_righe(p))

    rep = Report()
    controllo_1_tipo_effetto(righe, vocab, rep)
    controllo_2_ruolo(righe, vocab, rep)
    controllo_3_archetipo(righe, vocab, rep)
    controllo_4_range(righe, range_liv, rep, ERRORE if range_strict else AVVISO)
    controllo_5_codici_effetto(righe, rep)
    controllo_6_campi_stampa(righe, vocab, rep)
    controllo_7_keyword_orfane(righe, glossario, rep)
    controllo_8_doppia_presenza(righe, rep)
    come_misurato = controllo_9_budget_testo(righe, vocab, rep)

    return rep, {
        "file": [os.path.basename(p) for p in percorsi],
        "righe_totali": len(righe),
        "range_regolamento": range_liv,
        "budget_misurato_con": come_misurato,
    }


def stampa_report(rep, meta, solo_errori=False):
    print("=" * 78)
    print("WORLDLOOM — validazione dati carte")
    print("=" * 78)
    print(f"File analizzati : {', '.join(meta['file'])}")
    print(f"Righe carta     : {meta['righe_totali']}")
    print("Range cap. 8    : " + " · ".join(
        f"L{k}: A/P {v['attacco_parata'][0]}-{v['attacco_parata'][1]}, Vita {v['vita'][0]}-{v['vita'][1]}"
        for k, v in sorted(meta["range_regolamento"].items())))
    print(f"Budget testo    : misurato con {meta['budget_misurato_con']}")
    print()

    gruppi = {}
    for v in rep.voci:
        if solo_errori and v["livello"] != ERRORE:
            continue
        gruppi.setdefault((v["livello"], v["controllo"]), []).append(v)

    for (livello, controllo) in sorted(gruppi, key=lambda k: (k[0] != ERRORE, k[1])):
        voci = gruppi[(livello, controllo)]
        marcatore = "✗" if livello == ERRORE else "!"
        print(f"{marcatore} {livello} · {controllo} — {len(voci)}")
        for v in voci:
            posizione = f"{v['file']} · {v['foglio']} · r{v['riga']}" if v["riga"] else "(globale)"
            print(f"    {posizione}")
            print(f"      «{v['carta']}» — {v['messaggio']}")
        print()

    print("-" * 78)
    print(f"ERRORI: {len(rep.errori)}   AVVISI: {len(rep.avvisi)}")
    if rep.errori:
        print("ESITO: FALLITO — cards.json non va rigenerato finche' gli errori restano.")
    else:
        print("ESITO: OK — nessun errore bloccante.")
    print("-" * 78)


def main():
    ap = argparse.ArgumentParser(description="Validatore dei dati carte Worldloom")
    ap.add_argument("--solo-errori", action="store_true", help="nasconde gli avvisi nel report")
    ap.add_argument("--range-strict", action="store_true",
                    help="i range statistiche del cap. 8 diventano errori bloccanti")
    ap.add_argument("--json", action="store_true", help="stampa il report in JSON")
    args = ap.parse_args()

    rep, meta = valida(range_strict=args.range_strict)

    if args.json:
        print(json.dumps({"meta": {k: v for k, v in meta.items() if k != "range_regolamento"},
                          "voci": rep.voci,
                          "errori": len(rep.errori),
                          "avvisi": len(rep.avvisi)}, ensure_ascii=False, indent=2))
    else:
        stampa_report(rep, meta, solo_errori=args.solo_errori)

    return 1 if rep.errori else 0


if __name__ == "__main__":
    sys.exit(main())
